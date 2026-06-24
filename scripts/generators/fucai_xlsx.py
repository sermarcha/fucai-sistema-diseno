"""
fucai_xlsx.py — Reusable FUCAI Excel (.xlsx) styling helpers (openpyxl).

    import sys; sys.path.insert(0, "<skill>/scripts")
    from fucai_xlsx import apply_fucai_header, apply_fucai_body, style_title, RAMP_ORANGE

Minimalist by design: orange header, alternate arena rows, horizontal bottom
rule only (no full grid). Numbers right-aligned, units in the header.
Validate formulas with /mnt/skills/public/xlsx/scripts/recalc.py.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "lib"))
from tokens import C, FONT, RAMP_ORANGE, RAMP_GREEN, RAMP_NEUTRAL  # marca y rampas desde 03_tokens/tokens.json


HEADER_FILL = PatternFill(start_color=C["primary"], end_color=C["primary"], fill_type="solid")
HEADER_FONT = Font(name=FONT["body"], size=10, bold=True, color=C["white"])
BODY_FONT   = Font(name=FONT["body"], size=10, color=C["black"])
TITLE_FONT  = Font(name=FONT["heading"], size=14, bold=True, color=C["primary"])
ALT_FILL    = PatternFill(start_color=C["lightSand"], end_color=C["lightSand"], fill_type="solid")
BOTTOM_RULE = Border(bottom=Side(style="thin", color=C["grayLight"]))

# Rampas de visualización de datos (RAMP_ORANGE / RAMP_GREEN / RAMP_NEUTRAL): se
# importan ya resueltas desde 03_tokens/tokens.json (dataviz.ramp.*). Úsalas para
# gráficas; el verde SOLO para series de territorio/naturaleza.


def style_title(ws, cell="A1", text=None):
    if text is not None:
        ws[cell] = text
    ws[cell].font = TITLE_FONT


def apply_fucai_header(ws, row=1, start_col=1, end_col=None):
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_fucai_body(ws, start_row=2, end_row=None, start_col=1, end_col=None, numeric_from_col=2):
    """numeric_from_col: columns >= this index are right-aligned (numbers)."""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    for r in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            c = ws.cell(row=r, column=col)
            c.font = BODY_FONT
            c.border = BOTTOM_RULE
            c.alignment = Alignment(horizontal="right" if col >= numeric_from_col else "left", vertical="center")
            if (r - start_row) % 2 == 1:
                c.fill = ALT_FILL
