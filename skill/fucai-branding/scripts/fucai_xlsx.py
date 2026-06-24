"""
fucai_xlsx.py — Reusable FUCAI Excel (.xlsx) styling helpers (openpyxl).

    import sys; sys.path.insert(0, "<skill>/scripts")
    from fucai_xlsx import apply_fucai_header, apply_fucai_body, style_title, RAMP_ORANGE

Minimalist by design: orange header, alternate arena rows, horizontal bottom
rule only (no full grid). Numbers right-aligned, units in the header.
Validate formulas with /mnt/skills/public/xlsx/scripts/recalc.py.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

C = {"primary": "E94513", "arena": "EDE8D3", "green": "2D6A4F", "black": "000000",
     "white": "FFFFFF", "grayText": "333333", "grayLight": "CCCCCC", "lightSand": "F6F3E9"}

HEADER_FILL = PatternFill(start_color=C["primary"], end_color=C["primary"], fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=C["white"])
BODY_FONT   = Font(name="Calibri", size=10, color=C["black"])
TITLE_FONT  = Font(name="Space Grotesk", size=14, bold=True, color=C["primary"])
ALT_FILL    = PatternFill(start_color=C["lightSand"], end_color=C["lightSand"], fill_type="solid")
BOTTOM_RULE = Border(bottom=Side(style="thin", color=C["grayLight"]))

# Brand data ramps (use for charts; verde only for territory/nature series).
RAMP_ORANGE = ["C13A10", "E94513", "F06A3E", "F4A28A", "F8C4AE", "F6F3E9", "FBF9F3"]
RAMP_GREEN  = ["1B4032", "2D6A4F", "4E8A6F", "74B597", "A0D0B8", "C8E4D5", "EDE8D3"]
RAMP_NEUTRAL = ["000000", "333333", "666666", "999999", "CCCCCC", "E8E8E8", "FFFFFF"]


def style_title(ws, cell="A1", text=None):
    if text is not None:
        ws[cell] = text
    ws[cell].font = TITLE_FONT


def apply_fucai_header(ws, row=1, start_col=1, end_col=None):
    if end_col is None:
        end_col = ws.max_column
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_fucai_body(ws, start_row=2, end_row=None, start_col=1, end_col=None, numeric_from_col=2):
    """numeric_from_col: columns >= this index are right-aligned (numbers)."""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    for r in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            c = ws.cell(row=r, column=col)
            c.font = BODY_FONT
            c.border = BOTTOM_RULE
            c.alignment = Alignment(horizontal="right" if col >= numeric_from_col else "left", vertical="center")
            if (r - start_row) % 2 == 1:
                c.fill = ALT_FILL
