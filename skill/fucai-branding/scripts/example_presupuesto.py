"""
example_presupuesto.py — Plantilla de presupuesto FUCAI (.xlsx). Ejemplo ejecutable y base.

    python3 scripts/example_presupuesto.py [salida.xlsx]

Estructura genérica: logo + título, campo de proyecto en texto ABIERTO,
encabezado naranja, filas con fórmula (total = cantidad × valor unitario),
subtotal, formato COP. Reemplaza los datos de ejemplo. Valida fórmulas con
/mnt/skills/public/xlsx/scripts/recalc.py.
"""
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from fucai_xlsx import (apply_fucai_header, apply_fucai_body, TITLE_FONT, C)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

ASSETS = os.path.join(os.path.dirname(__file__), "..", "assets")
COP = '#,##0 "COP"'

def build(out):
    wb = Workbook(); ws = wb.active; ws.title = "Presupuesto"
    ws.sheet_view.showGridLines = False

    # Logo + título
    try:
        logo = XlImage(os.path.join(ASSETS, "logo_naranja.png")); logo.width = 151; logo.height = 76
        ws.add_image(logo, "A1")
    except Exception:
        pass
    ws.merge_cells("C1:F1"); ws["C1"] = "Presupuesto de Proyecto"; ws["C1"].font = TITLE_FONT
    ws.merge_cells("C2:F2"); ws["C2"] = "Nuestro centro es la periferia"
    ws["C2"].font = Font(name="Calibri", size=10, italic=True, color=C["primary"])

    # Campo de proyecto — TEXTO ABIERTO (nunca casilla fija)
    ws["A4"] = "Proyecto / Centro de costos:"; ws["A4"].font = Font(name="Calibri", size=10, bold=True, color=C["grayText"])
    ws.merge_cells("C4:F4"); ws["C4"] = "(escribir aquí — ej. CC217 Naane (OIKOS–AICS))"
    ws["C4"].font = Font(name="Calibri", size=10, color=C["grayText"])

    # Encabezado
    headers = ["Rubro", "Descripción", "Unidad", "Cantidad", "Valor unitario (COP)", "Valor total (COP)"]
    hrow = 6
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    apply_fucai_header(ws, row=hrow, start_col=1, end_col=6)

    # Filas de ejemplo (total con fórmula)
    rows = [
        ["B11 Talleres", "Taller de gobernanza territorial", "taller", 4, 1800000],
        ["B14 Transporte", "Transporte fluvial a comunidades", "viaje", 6, 950000],
        ["B17 Materiales", "Kit de semillas y herramientas", "kit", 14, 320000],
    ]
    start = hrow + 1
    for r, row in enumerate(rows, start=start):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
    end = start + len(rows) - 1
    apply_fucai_body(ws, start_row=start, end_row=end, start_col=1, end_col=6, numeric_from_col=4)

    # Formato COP en columnas de valor
    for r in range(start, end + 1):
        ws.cell(row=r, column=5).number_format = COP
        ws.cell(row=r, column=6).number_format = COP

    # Subtotal
    sub = end + 1
    ws.cell(row=sub, column=5, value="Subtotal").font = Font(name="Calibri", size=10, bold=True, color=C["primary"])
    ws.cell(row=sub, column=5).alignment = Alignment(horizontal="right")
    tot = ws.cell(row=sub, column=6, value=f"=SUM(F{start}:F{end})")
    tot.font = Font(name="Calibri", size=10, bold=True, color=C["primary"]); tot.number_format = COP
    tot.fill = PatternFill(start_color=C["lightSand"], end_color=C["lightSand"], fill_type="solid")
    tot.alignment = Alignment(horizontal="right")

    # Anchos de columna
    for col, w in zip("ABCDEF", [16, 34, 10, 10, 18, 18]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 60

    # Ajustar a una página de ancho al imprimir/exportar
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    wb.save(out); return out

if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/outputs/FUCAI_Presupuesto_Base_2026-06.xlsx"
    print("OK ->", build(out))
